from datetime import datetime
import io
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

st.set_page_config(
    page_title="å°ç£å¤§è»ŠéšŠå ±è¡¨æ•´ç† App",
    page_icon="ğŸ",
    layout="wide",
)


def parse_extension_input(extension_input):
    extension = {}
    for line in extension_input.split('\n'):
        parts = line.strip().split(':')
        if len(parts) == 2:
            employee_id, ext = parts
            extension[employee_id.strip()] = ext.strip()
    return extension


def process_dataframe(df):
    # æ‰¾åˆ° "æ—…æ¬¡æ˜ç´°è¡¨" æ‰€åœ¨çš„è¡Œ
    start_row = df[df.apply(lambda row: 'æ—…æ¬¡æ˜ç´°è¡¨' in str(row.values), axis=1)].index

    # æ‰¾åˆ° "ç¸½å…±ï¼š" æ‰€åœ¨çš„è¡Œ
    end_row = df[df.apply(lambda row: str(row.values[0]).startswith('ç¸½å…±ï¼š'), axis=1)].index

    # å°‹æ‰¾åˆ—å¸³æœŸé–“
    billing_period = ''
    for _, row in df.iterrows():
        if 'åˆ—å¸³æœŸé–“ï¼š' in str(row.values):
            billing_period = row.values[1]  # å‡è¨­åˆ—å¸³æœŸé–“åœ¨æœ€å¾Œä¸€åˆ—
            break

    if len(start_row) > 0 and len(end_row) > 0:
        # å¦‚æœæ‰¾åˆ° "æ—…æ¬¡æ˜ç´°è¡¨" å’Œ "ç¸½å…±ï¼š"ï¼Œå‰‡å–ä¸­é–“çš„æ•¸æ“š
        start_row = start_row[0] + 1
        end_row = end_row[0]
        df = df.iloc[start_row:end_row].reset_index(drop=True)

        # å°‡ç¬¬ä¸€è¡Œè¨­ç‚ºåˆ—æ¨™é¡Œ
        new_header = df.iloc[0]
        df = df[1:]
        df.columns = new_header

        # é‡ç½®ç´¢å¼•
        df = df.reset_index(drop=True)
    elif len(start_row) > 0:
        # å¦‚æœåªæ‰¾åˆ° "æ—…æ¬¡æ˜ç´°è¡¨"ï¼Œå‰‡å¾è©²è¡Œä¹‹å¾Œé–‹å§‹å–æ•¸æ“š
        start_row = start_row[0] + 1
        df = df.iloc[start_row:].reset_index(drop=True)

        # å°‡ç¬¬ä¸€è¡Œè¨­ç‚ºåˆ—æ¨™é¡Œ
        new_header = df.iloc[0]
        df = df[1:]
        df.columns = new_header

        # é‡ç½®ç´¢å¼•
        df = df.reset_index(drop=True)
    else:
        st.warning("æœªæ‰¾åˆ° 'æ—…æ¬¡æ˜ç´°è¡¨' è¡Œï¼Œé¡¯ç¤ºåŸå§‹æ•¸æ“šã€‚")

    return df, billing_period


def display_employee_data(df):
    employee_column = 'å“¡å·¥ç·¨è™Ÿ'
    name_column = 'å“¡å·¥å§“å'

    if employee_column not in df.columns and name_column not in df.columns:
        st.error(f"æ‰¾ä¸åˆ° '{employee_column}' æˆ– '{name_column}' åˆ—ã€‚è«‹ç¢ºä¿æ•¸æ“šä¸­åŒ…å«é€™äº›åˆ—ã€‚")
        return

    # ç²å–æ‰€æœ‰å”¯ä¸€çš„å“¡å·¥ç·¨è™Ÿå’Œå§“å
    employees = df[[employee_column, name_column]].drop_duplicates()

    # å‰µå»ºä¸€å€‹åŒ…å«å“¡å·¥ç·¨è™Ÿå’Œå§“åçš„é¸é …åˆ—è¡¨
    employee_options = [
        f"{row[employee_column]} - {row[name_column]}" for _, row in employees.iterrows()
    ]

    # å‰µå»ºä¸€å€‹é¸æ“‡æ¡†è®“ç”¨æˆ¶é¸æ“‡è¦æŸ¥çœ‹çš„å“¡å·¥
    selected_employee = st.selectbox("é¸æ“‡å“¡å·¥", employee_options)

    # å¾é¸æ“‡çš„é¸é …ä¸­æå“¡å·¥ç·¨è™Ÿ
    selected_employee_id = selected_employee.split(' - ')[0]

    # é¡¯ç¤ºé¸ä¸­å“¡å·¥çš„æ•¸æ“š
    employee_data = df[df[employee_column] == selected_employee_id]
    st.subheader(f"å“¡å·¥ {selected_employee} çš„æ•¸æ“š")
    st.dataframe(employee_data)


def create_employee_sheets(df, billing_period, original_file, grouped_employees, extension):
    employee_column = 'å“¡å·¥ç·¨è™Ÿ'
    name_column = 'å“¡å·¥å§“å'

    if employee_column not in df.columns or name_column not in df.columns:
        st.error(f"æ‰¾ä¸åˆ° '{employee_column}' æˆ– '{name_column}' åˆ—ã€‚è«‹ç¢ºä¿æ•¸æ“šä¸­åŒ…å«é€™äº›åˆ—ã€‚")
        return None

    # åŠ è¼‰åŸå§‹å·¥ä½œç°¿
    workbook = load_workbook(original_file)

    summary_sheet = workbook.create_sheet("ç¸½è¡¨")
    current_date = datetime.now()
    current_year_month = current_date.strftime("%Y/%m")
    today = current_date.strftime("%Y/%m/%d")
    fixed_rows = [
        ['å°ç£å¤§è»ŠéšŠä¹˜è»Šè²»ç¸½è¡¨', '', '', current_year_month],
        ['åˆ—å¸³æœŸé–“ï¼š', '', '', billing_period],
        ['æ”¶æ“šæ—¥æœŸ', '', '', today],
    ]

    for i, row in enumerate(fixed_rows):
        summary_sheet.append(row)
        summary_sheet.merge_cells(f'A{i+1}:C{i+1}')
        summary_sheet.merge_cells(f'D{i+1}:G{i+1}')

    summary_sheet.append(["NO", "å“¡å·¥å§“å", "å·¥è™Ÿ", "è¯çµ¡é›»è©±", "ç­†æ•¸", "æŠ˜æ‰£å¾Œè»Šè³‡", "ACK"])

    summary_data = []

    # è™•ç†åˆ†çµ„çš„å“¡å·¥
    for group, employees in grouped_employees.items():
        if not employees:
            continue

        # ä½¿ç”¨ grouped_employees ä¸­çš„ç¬¬ä¸€å€‹å“¡å·¥ç·¨è™Ÿä½œç‚ºä»£è¡¨
        first_employee_id = employees[0]
        group_df = df[df[employee_column].isin(employees)]

        if group_df.empty:
            continue

        # ç²å–ç¬¬ä¸€å€‹å“¡å·¥çš„è³‡è¨Š
        first_employee_data = group_df[group_df[employee_column] == first_employee_id].iloc[0]
        first_employee_name = first_employee_data[name_column]

        total_count = len(group_df)
        total_amount = group_df['æŠ˜æ‰£å¾Œè»Šè³‡'].sum() if 'æŠ˜æ‰£å¾Œè»Šè³‡' in group_df.columns else 0

        summary_data.append(
            [
                len(summary_data) + 1,
                first_employee_name,
                first_employee_id,
                extension.get(first_employee_id, ""),
                total_count,
                total_amount,
                "",
            ]
        )

        sheet_name = f'{first_employee_id} {first_employee_name}'
        worksheet = workbook.create_sheet(sheet_name)

        # åˆ›å»ºå›ºå®šçš„è¡Œå†…å®¹
        fixed_rows = [
            ['ä¼æ¥­æœƒå“¡ä¹˜è»Šæœå‹™é›»å­å°å¸³å–®'],
            ['å®¢æˆ¶åç¨±ï¼š', '', 'å‹è¨Šç§‘æŠ€è‚¡ä»½æœ‰é™å…¬å¸'],
            ['åˆ—å¸³æœŸé–“ï¼š', '', billing_period],
        ]

        # å°†å›ºå®šè¡Œå†…å®¹å†™å…¥å·¥ä½œè¡¨
        for row in fixed_rows:
            worksheet.append(row)

        # å°‡çµ„å…§æ‰€æœ‰å“¡å·¥çš„æ•¸æ“šå¯«å…¥å·¥ä½œè¡¨ï¼ŒæŒ‰ç…§å“¡å·¥ç·¨è™Ÿæ’åº
        sorted_group_df = group_df.sort_values(by=employee_column)
        worksheet.append(sorted_group_df.columns.tolist())
        for _, row in sorted_group_df.iterrows():
            worksheet.append(row.tolist())

        # è®¡ç®—ç»Ÿè®¡æ•°æ®
        total_count = len(sorted_group_df)
        total_amount = (
            sorted_group_df['æŠ˜æ‰£å¾Œè»Šè³‡'].sum() if 'æŠ˜æ‰£å¾Œè»Šè³‡' in sorted_group_df.columns else 0
        )

        # åˆ›å»ºç»Ÿè®¡æ•°æ®è¡Œ
        stats_rows = [
            ['ç¸½ç­†æ•¸', total_count, '', '', '', '', 'æŠ˜æ‰£å¾Œï¼š', total_amount],
            [],
            ['*è»Šè³‡ç¸½è¨ˆ(é‹é€æœå‹™è²»)ï¼š', '', '', '', '', '', f"{total_amount}å…ƒ"],
            ['ä¹˜è»Šåˆ¸å°è£½è²»ï¼š', '', '', '', '', '', '0å…ƒ'],
            ['æ»¯ç´é‡‘ï¼š', '', '', '', '', '', '0å…ƒ'],
            ['å…¶å®ƒè²»ç”¨ï¼š', '', '', '', '', '', '0å…ƒ'],
            ['æœ¬æœŸæ‡‰ç¹³å¸³æ¬¾ï¼š', '', '', '', '', '', f"{total_amount}å…ƒ"],
            ['ç‰¹æ®Šè²»ç”¨ï¼š', '', '', '', '', '', '0å…ƒ'],
        ]

        # å†™å…¥ç»Ÿè®¡æ•°æ®è¡Œ
        for row in stats_rows:
            worksheet.append(row)

        # è®¾ç½®å­—ä½“å¤§å°å’Œè°ƒæ•´åˆ—å®½
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = Font(size=12)

        for idx, column in enumerate(sorted_group_df.columns):
            column_letter = get_column_letter(idx + 1)
            if column in ['ä¸Šè»Šåœ°é»', 'ä¸‹è»Šåœ°é»']:
                worksheet.column_dimensions[column_letter].width = 12
            else:
                max_length = max(
                    sorted_group_df[column].astype(str).map(len).max() + 4, len(str(column)) + 6
                )
                worksheet.column_dimensions[column_letter].width = max_length

        # åˆä½µç¬¬ä¸€è¡Œå–®å…ƒæ ¼ä¸¦ç½®ä¸­
        max_col = len(sorted_group_df.columns)
        worksheet.merge_cells(f'A1:{get_column_letter(max_col)}1')
        title_cell = worksheet['A1']
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        worksheet.merge_cells(f'C2:{get_column_letter(max_col)}2')
        worksheet.merge_cells(f'C3:{get_column_letter(max_col)}3')

        # è®¾ç¬¬ä¸€è¡Œä¸ºç²—ä½“
        bold_font = Font(size=12, bold=True)
        title_cell.font = bold_font
        start_row = len(fixed_rows) + len(sorted_group_df) + 2
        total_count_cell = worksheet.cell(row=start_row, column=1)
        total_count_cell.font = bold_font
        total_count_cell = worksheet.cell(row=start_row, column=2)
        total_count_cell.font = bold_font
        total_amount_cell = worksheet.cell(row=start_row, column=7)
        total_amount_cell.font = bold_font
        total_amount_cell = worksheet.cell(row=start_row, column=8)
        total_amount_cell.font = bold_font
        payable_amount_cell = worksheet.cell(row=start_row + 6, column=1)
        payable_amount_cell.font = bold_font
        payable_amount_value_cell = worksheet.cell(row=start_row + 6, column=7)
        payable_amount_value_cell.font = bold_font

        # æ·»åŠ å¤–æ¡†çº¿
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # ä¸ºæ•°æ®éƒ¨åˆ†æ·»åŠ å…¨éƒ¨æ¡†çº¿
        for row in worksheet[f'A1':f'{get_column_letter(max_col)}{worksheet.max_row}']:
            for cell in row:
                cell.border = thin_border

    # è™•ç†æœªåˆ†çµ„çš„å“¡å·¥
    ungrouped_employees = set(df[employee_column]) - set(
        employee for group in grouped_employees.values() for employee in group
    )
    for employee in ungrouped_employees:
        employee_df = df[df[employee_column] == employee]
        employee_name = employee_df[name_column].iloc[0]

        total_count = len(employee_df)
        total_amount = employee_df['æŠ˜æ‰£å¾Œè»Šè³‡'].sum() if 'æŠ˜æ‰£å¾Œè»Šè³‡' in employee_df.columns else 0

        summary_data.append(
            [
                len(summary_data) + 1,
                employee_name,
                employee,
                extension.get(employee, ""),
                total_count,
                total_amount,
                "",
            ]
        )

        sheet_name = f'{employee} {employee_name}'
        worksheet = workbook.create_sheet(sheet_name)

        # åˆ›å»ºå›ºå®šçš„è¡Œå†…å®¹
        fixed_rows = [
            ['ä¼æ¥­æœƒå“¡ä¹˜è»Šæœå‹™é›»å­å°å¸³å–®'],
            ['å®¢æˆ¶åç¨±ï¼š', '', 'å‹è¨Šç§‘æŠ€è‚¡ä»½æœ‰é™å…¬å¸'],
            ['åˆ—å¸³æœŸé–“ï¼š', '', billing_period],
        ]

        # å°†å›ºå®šè¡Œå†…å®¹å†™å…¥å·¥ä½œè¡¨
        for row in fixed_rows:
            worksheet.append(row)

        # é‡ç½® employee_df çš„ç´¢å¼•å¹¶ä¿ç•™åŸå§‹åˆ—å
        employee_df_reset = employee_df.reset_index(drop=True)

        # å°†å‘˜å·¥æ•°æ®åŒ…æ‹¬æ ‡é¢˜ï¼‰å†™å…¥å·¥ä½œè¡¨
        worksheet.append(employee_df_reset.columns.tolist())
        for _, row in employee_df_reset.iterrows():
            worksheet.append(row.tolist())

        # è®¡ç®—ç»Ÿè®¡æ•°æ®
        total_count = len(employee_df_reset)
        total_amount = (
            employee_df_reset['æŠ˜æ‰£å¾Œè»Šè³‡'].sum()
            if 'æŠ˜æ‰£å¾Œè»Šè³‡' in employee_df_reset.columns
            else 0
        )

        # åˆ›å»ºç»Ÿè®¡æ•°æ®è¡Œ
        stats_rows = [
            ['ç¸½ç­†æ•¸', total_count, '', '', '', '', 'æŠ˜æ‰£å¾Œï¼š', total_amount],
            [],
            ['*è»Šè³‡ç¸½è¨ˆ(é‹é€æœå‹™è²»)ï¼š', '', '', '', '', '', f"{total_amount}å…ƒ"],
            ['ä¹˜è»Šåˆ¸å°è£½è²»ï¼š', '', '', '', '', '', '0å…ƒ'],
            ['æ»¯ç´é‡‘ï¼š', '', '', '', '', '', '0å…ƒ'],
            ['å…¶å®ƒè²»ç”¨ï¼š', '', '', '', '', '', '0å…ƒ'],
            ['æœ¬æœŸæ‡‰ç¹³å¸³æ¬¾ï¼š', '', '', '', '', '', f"{total_amount}å…ƒ"],
            ['ç‰¹æ®Šè²»ç”¨ï¼š', '', '', '', '', '', '0å…ƒ'],
        ]

        # å†™å…¥ç»Ÿè®¡æ•°æ®è¡Œ
        for row in stats_rows:
            worksheet.append(row)

        # è®¾ç½®å­—ä½“å¤§å°å’Œè°ƒæ•´åˆ—å®½
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = Font(size=12)

        for idx, column in enumerate(employee_df_reset.columns):
            column_letter = get_column_letter(idx + 1)
            if column in ['ä¸Šè»Šåœ°é»', 'ä¸‹è»Šåœ°é»']:
                worksheet.column_dimensions[column_letter].width = 12
            else:
                max_length = max(
                    employee_df_reset[column].astype(str).map(len).max() + 4, len(str(column)) + 6
                )
                worksheet.column_dimensions[column_letter].width = max_length

        # åˆä½µç¬¬ä¸€è¡Œå–®å…ƒæ ¼ä¸¦ç½®ä¸­
        max_col = len(employee_df_reset.columns)
        worksheet.merge_cells(f'A1:{get_column_letter(max_col)}1')
        title_cell = worksheet['A1']
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        worksheet.merge_cells(f'C2:{get_column_letter(max_col)}2')
        worksheet.merge_cells(f'C3:{get_column_letter(max_col)}3')

        # è®¾ç½®ç¬¬ä¸€è¡Œä¸ºç²—ä½“
        bold_font = Font(size=12, bold=True)
        title_cell.font = bold_font
        start_row = len(fixed_rows) + len(employee_df_reset) + 2
        total_count_cell = worksheet.cell(row=start_row, column=1)
        total_count_cell.font = bold_font
        total_count_cell = worksheet.cell(row=start_row, column=2)
        total_count_cell.font = bold_font
        total_amount_cell = worksheet.cell(row=start_row, column=7)
        total_amount_cell.font = bold_font
        total_amount_cell = worksheet.cell(row=start_row, column=8)
        total_amount_cell.font = bold_font
        payable_amount_cell = worksheet.cell(row=start_row + 6, column=1)
        payable_amount_cell.font = bold_font
        payable_amount_value_cell = worksheet.cell(row=start_row + 6, column=7)
        payable_amount_value_cell.font = bold_font

        # æ·»åŠ å¤–æ¡†çº¿
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # ä¸ºæ•°æ®éƒ¨åˆ†æ·»åŠ å…¨éƒ¨æ¡†çº¿
        for row in worksheet[f'A1':f'{get_column_letter(max_col)}{worksheet.max_row}']:
            for cell in row:
                cell.border = thin_border

    # æ ¹æ“šå“¡å·¥ç·¨è™Ÿæ’åº summary_data
    summary_data.sort(key=lambda x: x[2])  # x[2] æ˜¯å“¡å·¥ç·¨è™Ÿ

    # å°‡æ’åºå¾Œçš„ summary_data å¯«å…¥ç¸½è¡¨
    for i, row in enumerate(summary_data, start=1):
        row[0] = i  # æ›´æ–°åºè™Ÿ
        summary_sheet.append(row)

    # è®¡ç®—æ€»è®¡
    total_count = sum(row[4] for row in summary_data)
    total_amount = sum(row[5] for row in summary_data)
    summary_sheet.append(["åˆè¨ˆ", "", "", "", total_count, total_amount, ""])
    summary_sheet.merge_cells(f"A{len(summary_data) + 5}:D{len(summary_data) + 5}")

    # è®¾ç½®æ€»è¡¨æ ¼å¼
    for row in summary_sheet.iter_rows(
        min_row=1, max_row=len(summary_data) + 5, min_col=1, max_col=7
    ):
        for cell in row:
            cell.font = Font(size=12)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'),
            )

    for row in summary_sheet.iter_rows(min_row=1, max_row=3):
        for col, cell in enumerate(row, start=1):
            if col == 1:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 4:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    for row in summary_sheet.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # åœ¨è™•ç†å®Œæ‰€æœ‰å·¥ä½œè¡¨å¾Œï¼Œé‡æ–°æ’åº
    sheets = workbook.sheetnames
    # åªå°å“¡å·¥å·¥ä½œè¡¨é€²è¡Œæ’åºï¼Œä¿ç•™å‰å…©å€‹å·¥ä½œè¡¨ä¸è®Š
    employee_sheets = sheets[2:]

    # æ ¹æ“šå·¥ä½œè¡¨åç¨±ï¼ˆå“¡å·¥ç·¨è™Ÿï¼‰é€²è¡Œæ’åº
    sorted_sheets = sorted(employee_sheets, key=lambda x: x.split()[0])

    # é‡æ–°æ’åˆ—å·¥ä½œè¡¨
    for i, sheet_name in enumerate(sorted_sheets, start=2):
        workbook.move_sheet(sheet_name, offset=i - workbook.index(workbook[sheet_name]))

    # å°†ä¿®æ”¹åçš„å·¥ä½œä¿å­˜åˆ°å†…å­˜ä¸­
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


def get_all_employee_ids(df):
    employee_column = 'å“¡å·¥ç·¨è™Ÿ'
    if employee_column not in df.columns:
        st.error(f"æ‰¾ä¸åˆ° '{employee_column}' åˆ—ã€‚è«‹ç¢ºä¿æ•¸æ“šä¸­åŒ…å«æ­¤åˆ—ã€‚")
        return []
    return sorted(df[employee_column].unique().tolist())


def main():
    st.title("Excelæ•¸æ“šæ•´ç†å·¥å…·")

    # ä¸Šä¼ Excelæ–‡ä»¶
    uploaded_file = st.file_uploader("è«‹ä¸Šå‚³Excelæ–‡ä»¶", type=["xlsx", "xls"])

    if uploaded_file is not None:
        # ç²å–ä¸Šå‚³æ–‡ä»¶çš„åŸå§‹åç¨±
        original_filename = uploaded_file.name

        # è®€å–æ‰€æœ‰å·¥ä½œè¡¨
        xls = pd.ExcelFile(uploaded_file)
        sheet_names = xls.sheet_names

        # è®“ç”¨æˆ¶é¸æ“‡å·¥ä½œè¡¨
        selected_sheet = st.selectbox("è«‹é¸æ“‡è¦è™•ç†çš„å·¥ä½œè¡¨", sheet_names)

        # è®€å–é¸å®šçš„å·¥ä½œè¡¨
        df = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=None)

        # é¡¯ç¤ºåŸå§‹æ•¸æ“š
        st.subheader(f"åŸå§‹æ•¸æ“š - {selected_sheet}")
        st.dataframe(df)

        # è™•ç†æ•¸æ“š
        processed_df, billing_period = process_dataframe(df)

        # é¡¯ç¤ºè™•ç†å¾Œçš„æ•¸æ“š
        st.subheader(f"è™•ç†å¾Œçš„æ•¸æ“š - {selected_sheet}")
        st.dataframe(processed_df)

        # é¡¯ç¤ºæ¯å€‹å“¡å·¥çš„æ•¸æ“š
        display_employee_data(processed_df)

        # è·å–æ‰€æœ‰å‘˜å·¥ç¼–å·
        all_employee_ids = get_all_employee_ids(processed_df)

        # æ·»åŠ è¾“å…¥æ¡†è®©ç”¨æˆ·è¾“å…¥å‘˜å·¥ç¼–å·å’Œåˆ†æœºçš„å¯¹åº”å…³ç³»
        default_extension_input = "08956: 6312\n07030: 6412\n05259: 2340\n06294: 8254\n08332: 6654\n09025: 6716\n09092: 6112\n09137: 6834\n09214: 5738\n09324: 2531\n07468: 6417\n08951: 6300\n09021: 6413\n09335: 6416"
        extension_input = st.text_area(
            "è«‹è¼¸å…¥å“¡å·¥ç·¨è™Ÿå’Œåˆ†æ©Ÿçš„å°æ‡‰é—œä¿‚ï¼ˆæ¯è¡Œä¸€å€‹ï¼Œæ ¼å¼ç‚º 'å“¡å·¥ç·¨è™Ÿ: åˆ†æ©Ÿ'ï¼‰",
            default_extension_input,
        )

        # è§£æç”¨æˆ·è¾“å…¥çš„åˆ†æœºå¯¹åº”å…³ç³»
        extension = parse_extension_input(extension_input)

        # æ‰¾å‡ºç¼ºæ¼çš„å‘˜å·¥ç¼–å·
        missing_ids = [emp_id for emp_id in all_employee_ids if emp_id not in extension]

        # å¦‚æœæœ‰ç¼ºæ¼çš„å‘˜å·¥ç¼–å·ï¼Œæ›´æ–°è¾“å…¥æ¡†
        if missing_ids:
            missing_input = "\n".join([f"{emp_id}: " for emp_id in missing_ids])
            updated_extension_input = missing_input + "\n" + extension_input
            st.warning(f"ä»¥ä¸‹å“¡å·¥ç·¨è™Ÿç¼ºå°‘åˆ†æ©Ÿå°æ‡‰é—œä¿‚ï¼š{', '.join(missing_ids)}")
            extension_input = st.text_area(
                "æ›´æ–°å¾Œçš„å“¡å·¥ç·¨è™Ÿå’Œåˆ†æ©Ÿå°æ‡‰é—œä¿‚ï¼ˆè«‹ç‚ºç¼ºæ¼çš„ç·¨è™Ÿæ·»åŠ åˆ†æ©Ÿï¼‰",
                updated_extension_input,
            )
            extension = parse_extension_input(extension_input)

        # æ·»åŠ è¾“å…¥æ¡†è®©ç”¨æˆ·è¾“å…¥è¦åˆ†ç»„çš„å‘˜å·¥ç¼–å·
        grouped_employees_input = st.text_area(
            "è«‹è¼¸å…¥è¦ä¸€èµ·åˆ†çµ„çš„å“¡å·¥ç·¨è™Ÿï¼ˆæ¯çµ„ä¸€è¡Œï¼Œç”¨é€—è™Ÿåˆ†éš”ï¼‰",
            "",
        )

        # è™•ç†ç”¨æˆ¶è¼¸å…¥çš„åˆ†çµ„ä¿¡æ¯
        grouped_employees = {}
        if grouped_employees_input.strip():  # åªæœ‰ç•¶è¼¸å…¥ä¸ç‚ºç©ºæ™‚æ‰è™•ç†
            for i, group in enumerate(grouped_employees_input.split('\n')):
                employees = [emp.strip() for emp in group.split(',') if emp.strip()]
                if employees:
                    grouped_employees[f'Group_{i+1}'] = employees

        # å‰µå»ºåŒ…å«æ¯å€‹å“¡å·¥æ•¸æ“šçš„Excelæ–‡ä»¶
        output = create_employee_sheets(
            processed_df, billing_period, uploaded_file, grouped_employees, extension
        )

        if output:
            # æä¾›ä¸‹è½½æŒ‰é’®ï¼Œä½¿ç”¨åŸå§‹æ–‡ä»¶å
            st.download_button(
                label="ä¸‹è¼‰ä¿®æ”¹å¾Œçš„Excelæ–‡ä»¶",
                data=output.getvalue(),
                file_name=original_filename.replace('.xlsx', '_æ›´æ–°.xlsx'),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


if __name__ == "__main__":
    main()
